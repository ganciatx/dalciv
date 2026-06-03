#!/usr/bin/env python3
"""
Merge Revenue_Budget RevenueSources sheet into revsource-type-map.json.
Usage: python scripts/sync_budget_revsource_map.py [path/to/Revenue_Budget_*.xlsx]
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("pip install openpyxl") from None

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "dashboard/static/dallas-budget/revsource-type-map.json"
DISPLAY_OUT = ROOT / "dashboard/static/dallas-budget/revsource-display-map.json"
DEFAULT_XLSX = Path.home() / "Downloads/Revenue_Budget_20260520.xlsx"


def load_sheet_mapping(xlsx: Path) -> tuple[dict[str, str], dict[str, str]]:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if "RevenueSources" not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"No RevenueSources sheet in {xlsx}")
    ws = wb["RevenueSources"]
    mapping: dict[str, str] = {}
    display: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        revtype = row[7] if len(row) > 7 else None
        if not revtype:
            continue
        revtype = str(revtype).strip()
        plain = row[6]
        plain_label = str(plain).strip() if plain else ""
        for key in (row[2], row[6], row[0]):
            if key:
                mapping[str(key).strip()] = revtype
        if plain_label:
            for key in (row[0], row[2]):
                if key:
                    display[str(key).strip()] = plain_label
    wb.close()
    return mapping, display


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.is_file():
        raise SystemExit(f"File not found: {xlsx}")

    sheet_map, sheet_display = load_sheet_mapping(xlsx)
    existing: dict[str, str] = {}
    if OUT.is_file():
        existing = json.loads(OUT.read_text(encoding="utf-8"))

    merged = {**existing, **sheet_map}
    OUT.write_text(json.dumps(merged, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(f"Wrote {len(merged)} entries to {OUT} (+{len(sheet_map)} from spreadsheet)")

    display_existing: dict[str, str] = {}
    if DISPLAY_OUT.is_file():
        display_existing = json.loads(DISPLAY_OUT.read_text(encoding="utf-8"))
    display_merged = {**display_existing, **sheet_display}
    DISPLAY_OUT.write_text(
        json.dumps(display_merged, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {len(display_merged)} display labels to {DISPLAY_OUT}")


if __name__ == "__main__":
    main()
