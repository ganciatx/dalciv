#!/usr/bin/env python3
"""
Build revsource-display-map.json (cryptic revsource / rsrc code → plain-language label).

Prefer Revenue_Budget RevenueSources sheet (column G plain language) when an xlsx is available;
otherwise humanize revsource strings from the revenue cache.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "dashboard/static/dallas-budget/revsource-display-map.json"
CACHE = ROOT / "scraper_dashboard_data/revenue_budget_cache.json"
DEFAULT_XLSX = Path.home() / "Downloads/Revenue_Budget_20260520.xlsx"


def humanize_revsource(raw: str) -> str:
    s = (raw or "").strip()
    if not s:
        return s
    if re.match(r"^[A-Z][a-z]", s) and " " in s and not re.search(r"\.[A-Z]", s):
        return s
    s = re.sub(r"^Prop\.Taxs-", "Property taxes — ", s, flags=re.I)
    s = re.sub(r"^Chgs Serv-", "Service charges — ", s, flags=re.I)
    s = re.sub(r"^Intfd-", "Interfund — ", s, flags=re.I)
    s = re.sub(r"^Fines/For-", "Fines and forfeitures — ", s, flags=re.I)
    s = re.sub(r"^Fines/Fro-", "Fines — ", s, flags=re.I)
    s = re.sub(r"^Taxes-", "Taxes — ", s, flags=re.I)
    s = re.sub(r"Real Est", "real estate", s, flags=re.I)
    s = re.sub(r"Rl Est", "real estate", s, flags=re.I)
    s = re.sub(r"\bTxs\b", "taxes", s, flags=re.I)
    s = re.sub(r"Est-", "estate — ", s, flags=re.I)
    s = re.sub(r"\bRev\b", "revenue", s, flags=re.I)
    s = re.sub(r"-", " · ", s)
    return re.sub(r"\s+", " ", s).strip()


def load_from_xlsx(xlsx: Path) -> dict[str, str]:
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("pip install openpyxl") from None

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if "RevenueSources" not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"No RevenueSources sheet in {xlsx}")
    ws = wb["RevenueSources"]
    display: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 7:
            continue
        plain = row[6]
        if not plain:
            continue
        label = str(plain).strip()
        for key in (row[0], row[2]):
            if key:
                display[str(key).strip()] = label
    wb.close()
    return display


def load_from_cache() -> dict[str, str]:
    if not CACHE.is_file():
        return {}
    rows = json.loads(CACHE.read_text(encoding="utf-8")).get("rows") or []
    display: dict[str, str] = {}
    for row in rows:
        src = str(row.get("revsource") or "").strip()
        code = str(row.get("rsrc") or "").strip()
        pl = str(row.get("revsource_pl") or "").strip()
        label = pl or humanize_revsource(src)
        if src and label:
            display[src] = label
        if code and label:
            display[code] = label
    return display


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if xlsx.is_file():
        display = load_from_xlsx(xlsx)
        source = f"xlsx:{xlsx.name}"
    else:
        display = load_from_cache()
        source = "cache+humanize"

    existing: dict[str, str] = {}
    if OUT.is_file():
        existing = json.loads(OUT.read_text(encoding="utf-8"))
    merged = {**existing, **display}
    OUT.write_text(json.dumps(merged, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(f"Wrote {len(merged)} display labels to {OUT} ({source})")


if __name__ == "__main__":
    main()
